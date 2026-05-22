"""
新媒体Excel生成工具脚本
用法:
    python make_excel.py --brand trumpchi --type weekly --output weekly.xlsx
    python make_excel.py --brand audi --type kpi --data data.json --output kpi.xlsx
"""

import argparse
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation


# ============================================================
# 品牌样式
# ============================================================
BRAND_STYLES = {
    "trumpchi": {
        "header_fill": PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid"),
        "alt_fill": PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"),
    },
    "audi": {
        "header_fill": PatternFill(start_color="000000", end_color="000000", fill_type="solid"),
        "alt_fill": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    },
    "hyper": {
        "header_fill": PatternFill(start_color="6A0DAD", end_color="6A0DAD", fill_type="solid"),
        "alt_fill": PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),
    },
    "aion": {
        "header_fill": PatternFill(start_color="00A86B", end_color="00A86B", fill_type="solid"),
        "alt_fill": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    },
}

HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

FMT_MONEY = '#,##0.00'
FMT_PCT = '0.0%'
FMT_INT = '#,##0'
FMT_DATE = 'YYYY-MM-DD'


# ============================================================
# 通用工具
# ============================================================
def _style_header(ws, brand, headers, row=1):
    styles = BRAND_STYLES[brand]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = styles["header_fill"]
        c.font = HEADER_FONT
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def _write_rows(ws, brand, rows, start_row=2):
    styles = BRAND_STYLES[brand]
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            c.border = THIN_BORDER
            c.alignment = CENTER_ALIGN
            if ri % 2 == 1:
                c.fill = styles["alt_fill"]


def _fmt_col(ws, col_letter, start, end, fmt):
    for r in range(start, end + 1):
        ws[f"{col_letter}{r}"].number_format = fmt


def _finalize(ws, freeze="A2"):
    ws.freeze_panes = freeze
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    # 自适应列宽
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max_len + 4, 45)


# ============================================================
# 各类型生成
# ============================================================
def generate_weekly(brand, data, output_path):
    """周报数据表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "周报数据"

    headers = ["平台", "粉丝数", "本周新增", "发布量", "曝光量", "互动量", "互动率", "转化数", "转化率", "环比趋势"]
    _style_header(ws, brand, headers)
    _write_rows(ws, brand, data.get("rows", []))

    last_row = len(data.get("rows", [])) + 1
    _fmt_col(ws, "G", 2, last_row, FMT_PCT)
    _fmt_col(ws, "I", 2, last_row, FMT_PCT)

    ws.conditional_formatting.add(
        f"J2:J{last_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                   font=Font(color="006100"))
    )

    _finalize(ws)
    wb.save(output_path)
    return output_path


def generate_monthly(brand, data, output_path):
    """月度汇总表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "月度汇总"

    headers = ["平台", "月初粉丝", "月末粉丝", "净增", "总发布量", "总曝光", "总互动", "平均互动率", "总线索", "线索成本", "KPI达成率"]
    _style_header(ws, brand, headers)
    _write_rows(ws, brand, data.get("rows", []))

    last_row = len(data.get("rows", [])) + 1
    _fmt_col(ws, "H", 2, last_row, FMT_PCT)
    _fmt_col(ws, "J", 2, last_row, FMT_MONEY)
    _fmt_col(ws, "K", 2, last_row, FMT_PCT)

    ws.conditional_formatting.add(
        f"K2:K{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"K2:K{last_row}",
        CellIsRule(operator="lessThan", formula=["0.6"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    _finalize(ws)
    wb.save(output_path)
    return output_path


def generate_calendar(brand, data, output_path):
    """内容排期表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "内容排期"

    headers = ["日期", "平台", "内容标题", "内容类型", "目标人群", "发布时间", "状态", "负责人", "备注"]
    _style_header(ws, brand, headers)
    _write_rows(ws, brand, data.get("rows", []))

    last_row = len(data.get("rows", [])) + 1
    dv = DataValidation(type="list", formula1='"待发布,已发布,制作中,审核中"', allow_blank=True)
    dv.add(f"G2:G{last_row}")
    ws.add_data_validation(dv)

    _finalize(ws)
    wb.save(output_path)
    return output_path


def generate_kpi(brand, data, output_path):
    """KPI追踪表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI追踪"

    headers = ["指标", "年度目标", "本月目标", "当前完成", "完成率", "剩余", "日均需完成", "趋势", "状态"]
    _style_header(ws, brand, headers)
    _write_rows(ws, brand, data.get("rows", []))

    last_row = len(data.get("rows", [])) + 1
    _fmt_col(ws, "E", 2, last_row, FMT_PCT)

    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="greaterThan", formula=["1"],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                   font=Font(bold=True, color="006100"))
    )
    ws.conditional_formatting.add(
        f"E2:E{last_row}",
        CellIsRule(operator="lessThan", formula=["0.6"],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                   font=Font(bold=True, color="9C0006"))
    )

    _finalize(ws)
    wb.save(output_path)
    return output_path


def generate_competitor(brand, data, output_path):
    """竞品对比表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "竞品对比"

    headers = ["品牌", "平台", "粉丝数", "月发布量", "月曝光量", "月互动量", "互动率", "月线索数", "预估投放费用", "综合评分"]
    _style_header(ws, brand, headers)
    _write_rows(ws, brand, data.get("rows", []))

    last_row = len(data.get("rows", [])) + 1
    _fmt_col(ws, "G", 2, last_row, FMT_PCT)
    _fmt_col(ws, "I", 2, last_row, FMT_MONEY)

    ws.conditional_formatting.add(
        f"J2:J{last_row}",
        CellIsRule(operator="greaterThan", formula=["80"],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    _finalize(ws)
    wb.save(output_path)
    return output_path


# ============================================================
# 主流程
# ============================================================
GENERATORS = {
    "weekly": generate_weekly,
    "monthly": generate_monthly,
    "calendar": generate_calendar,
    "kpi": generate_kpi,
    "competitor": generate_competitor,
}


def main():
    parser = argparse.ArgumentParser(description="新媒体Excel生成工具")
    parser.add_argument("--brand", required=True,
                        choices=["trumpchi", "audi", "hyper", "aion"],
                        help="品牌标识")
    parser.add_argument("--type", dest="excel_type", default="weekly",
                        choices=list(GENERATORS.keys()),
                        help="表格类型")
    parser.add_argument("--data", help="JSON数据文件路径")
    parser.add_argument("--output", required=True, help="输出Excel文件路径")
    args = parser.parse_args()

    if args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = {"rows": []}

    gen = GENERATORS[args.excel_type]
    output = gen(args.brand, data, args.output)
    print(f"Excel 已生成: {output}")


if __name__ == "__main__":
    main()
